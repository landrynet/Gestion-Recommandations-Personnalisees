import json
from decimal import Decimal, InvalidOperation

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST

from .models import Note
from .forms import NoteForm
from subjects.models import MatiereClasse
from students.models import Student
from classes.models import Classe
from school_settings.models import SchoolInfo


# ─── Saisie des notes ─────────────────────────────────────────────────────────
@login_required
def saisie_notes(request):
    """Saisie des notes — enseignant (ses classes), préfet interdit ici."""
    user = request.user

    if user.is_prefet():
        messages.info(request, "La saisie des notes est réservée aux enseignants.")
        return redirect('consulter_notes')

    periodes = [
        ('1P',        '1ère Période'),
        ('2P',        '2ème Période'),
        ('EXAM1',     'Examen S1'),
        ('3P',        '3ème Période'),
        ('4P',        '4ème Période'),
        ('EXAM2',     'Examen S2'),
        ('REPECHAGE', 'Repêchage'),
    ]

    try:
        teacher = user.teacher_profile
        matieres_classes = MatiereClasse.objects.filter(
            enseignant=teacher
        ).select_related('matiere', 'classe', 'classe__section')
    except Exception:
        matieres_classes = MatiereClasse.objects.none()

    mc_id          = request.GET.get('mc', '')
    periode        = request.GET.get('periode', '1P')
    matiere_classe = None
    eleves         = []
    form           = None

    if mc_id:
        matiere_classe = get_object_or_404(MatiereClasse, pk=mc_id)
        try:
            if matiere_classe.enseignant != user.teacher_profile:
                messages.error(request, "Vous ne pouvez modifier que vos propres matières.")
                return redirect('saisie_notes')
        except Exception:
            messages.error(request, "Profil enseignant introuvable.")
            return redirect('dashboard')

        eleves = Student.objects.filter(classe=matiere_classe.classe).order_by('nom', 'postnom')

        if request.method == 'POST':
            form = NoteForm(request.POST, eleves=eleves, matiere_classe=matiere_classe, periode=periode)
            if form.is_valid():
                form.save()
                messages.success(request, f"Notes enregistrées — {matiere_classe.matiere} / {periode}.")
                return redirect(f'/notes/?mc={mc_id}&periode={periode}')
        else:
            form = NoteForm(eleves=eleves, matiere_classe=matiere_classe, periode=periode)

    return render(request, 'grades/saisie_notes.html', {
        'matieres_classes': matieres_classes,
        'mc_id':            mc_id,
        'periodes':         periodes,
        'periode':          periode,
        'matiere_classe':   matiere_classe,
        'eleves':           eleves,
        'form':             form,
    })


# ─── Auto-save AJAX ───────────────────────────────────────────────────────────
@login_required
@require_POST
def autosave_note(request):
    """Sauvegarde automatique d'une note individuelle via AJAX."""
    try:
        mc_id    = request.POST.get('mc_id')
        eleve_id = request.POST.get('eleve_id')
        periode  = request.POST.get('periode')
        valeur   = request.POST.get('valeur', '').strip()

        if not all([mc_id, eleve_id, periode]):
            return JsonResponse({'success': False, 'error': 'Données manquantes'}, status=400)

        mc    = get_object_or_404(MatiereClasse, pk=mc_id)
        eleve = get_object_or_404(Student, pk=eleve_id)

        # Vérification sécurité : l'enseignant ne peut sauver que ses matières
        user = request.user
        if not user.is_prefet():
            try:
                if mc.enseignant != user.teacher_profile:
                    return JsonResponse({'success': False, 'error': 'Non autorisé'}, status=403)
            except Exception:
                return JsonResponse({'success': False, 'error': 'Profil introuvable'}, status=403)

        # Calcul du maxima selon la période
        max_val = mc.matiere.maxima * (2 if periode in ('EXAM1', 'EXAM2') else 1)

        if valeur == '' or valeur is None:
            # Supprimer la note si le champ est vidé
            Note.objects.filter(eleve=eleve, matiere_classe=mc, periode=periode).delete()
            return JsonResponse({'success': True, 'action': 'deleted'})

        try:
            val_decimal = Decimal(valeur.replace(',', '.'))
        except InvalidOperation:
            return JsonResponse({'success': False, 'error': 'Valeur invalide'}, status=400)

        if val_decimal < 0 or val_decimal > max_val:
            return JsonResponse({
                'success': False,
                'error': f'La note doit être entre 0 et {max_val}'
            }, status=400)

        Note.objects.update_or_create(
            eleve=eleve,
            matiere_classe=mc,
            periode=periode,
            defaults={'valeur': val_decimal}
        )
        return JsonResponse({'success': True, 'action': 'saved', 'valeur': str(val_decimal)})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ─── Consultation des notes ───────────────────────────────────────────────────
@login_required
def consulter_notes(request):
    """Consultation des notes — préfet voit tout, enseignant voit ses classes."""
    user = request.user
    periodes = [
        ('1P', '1ère P'), ('2P', '2ème P'), ('EXAM1', 'Exam S1'),
        ('3P', '3ème P'), ('4P', '4ème P'), ('EXAM2', 'Exam S2'),
        ('REPECHAGE', 'Repêchage'),
    ]

    if user.is_prefet():
        classes = Classe.objects.select_related('section', 'annee_scolaire')
    else:
        try:
            teacher = user.teacher_profile
            mc_qs   = MatiereClasse.objects.filter(enseignant=teacher).values_list('classe_id', flat=True)
            classes = Classe.objects.filter(pk__in=mc_qs).select_related('section')
        except Exception:
            classes = Classe.objects.none()

    classe_id        = request.GET.get('classe', '')
    mc_id            = request.GET.get('mc', '')
    matieres_classes = []
    eleves_notes     = []
    selected_mc      = None

    if classe_id:
        if user.is_prefet():
            matieres_classes = MatiereClasse.objects.filter(
                classe_id=classe_id
            ).select_related('matiere', 'enseignant', 'enseignant__user').order_by('matiere__maxima', 'matiere__nom')
        else:
            try:
                matieres_classes = MatiereClasse.objects.filter(
                    classe_id=classe_id, enseignant=user.teacher_profile
                ).select_related('matiere', 'enseignant', 'enseignant__user')
            except Exception:
                matieres_classes = []

    if mc_id:
        selected_mc = get_object_or_404(MatiereClasse, pk=mc_id)
        eleves = Student.objects.filter(classe=selected_mc.classe).order_by('nom', 'postnom')
        for eleve in eleves:
            row = {'eleve': eleve, 'notes': {}}
            for code, label in periodes:
                try:
                    n = Note.objects.get(eleve=eleve, matiere_classe=selected_mc, periode=code)
                    row['notes'][code] = n.valeur
                except Note.DoesNotExist:
                    row['notes'][code] = None
            eleves_notes.append(row)

    return render(request, 'grades/consulter_notes.html', {
        'classes':          classes,
        'classe_id':        classe_id,
        'mc_id':            mc_id,
        'matieres_classes': matieres_classes,
        'selected_mc':      selected_mc,
        'eleves_notes':     eleves_notes,
        'periodes':         periodes,
    })


# ─── Export Excel ─────────────────────────────────────────────────────────────
@login_required
def export_notes_excel(request):
    """Exporte un fichier Excel vierge (ou pré-rempli) pour saisie hors-ligne."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        messages.error(request, "Module openpyxl manquant. Contactez l'administrateur.")
        return redirect('saisie_notes')

    mc_id   = request.GET.get('mc', '')
    periodes_sel = request.GET.getlist('periodes')

    if not mc_id:
        messages.error(request, "Veuillez sélectionner une matière/classe.")
        return redirect('saisie_notes')

    mc = get_object_or_404(MatiereClasse, pk=mc_id)

    # Sécurité
    if not request.user.is_prefet():
        try:
            if mc.enseignant != request.user.teacher_profile:
                messages.error(request, "Non autorisé.")
                return redirect('saisie_notes')
        except Exception:
            return redirect('saisie_notes')

    all_periodes = [
        ('1P', '1ère Période'),  ('2P', '2ème Période'),
        ('EXAM1', 'Examen S1'),  ('3P', '3ème Période'),
        ('4P', '4ème Période'),  ('EXAM2', 'Examen S2'),
        ('REPECHAGE', 'Repêchage'),
    ]
    if periodes_sel:
        periodes = [(c, l) for c, l in all_periodes if c in periodes_sel]
    else:
        periodes = all_periodes

    eleves = Student.objects.filter(classe=mc.classe).order_by('nom', 'postnom')
    school = SchoolInfo.get_info()

    nb_cols = 4 + len(periodes)  # N°, Matricule, Nom, Postnom + périodes
    last_col_letter = get_column_letter(nb_cols)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{mc.matiere.nom[:28]}"

    max_normal = mc.matiere.maxima

    # ── Styles réutilisables ──
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left      = Alignment(horizontal='left',   vertical='center')
    right_al  = Alignment(horizontal='right',  vertical='center')

    def thin_border(color='BDBDBD'):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def thick_border():
        tk = Side(style='medium', color='1565C0')
        return Border(left=tk, right=tk, top=tk, bottom=tk)

    fill_blue_dark  = PatternFill('solid', fgColor='1565C0')   # en-tête colonnes
    fill_blue_mid   = PatternFill('solid', fgColor='1976D2')   # sous-en-tête
    fill_blue_light = PatternFill('solid', fgColor='E3F2FD')   # maxima
    fill_gray_light = PatternFill('solid', fgColor='F5F5F5')   # colonnes fixes élèves
    fill_stripe     = PatternFill('solid', fgColor='FAFAFA')   # lignes paires
    fill_note       = PatternFill('solid', fgColor='FFFFFF')   # cellules notes

    white_bold = Font(bold=True, color='FFFFFF', size=10)
    white_reg  = Font(color='FFFFFF', size=9)
    dark_bold  = Font(bold=True, color='212121', size=10)
    dark_reg   = Font(color='424242', size=9)
    muted_font = Font(color='757575', size=8, italic=True)
    blue_font  = Font(bold=True, color='1565C0', size=11)

    # ══════════════════════════════════════════════
    # LIGNE 1 : République Démocratique du Congo
    # ══════════════════════════════════════════════
    ws.merge_cells(f'A1:{last_col_letter}1')
    c = ws['A1']
    c.value     = 'RÉPUBLIQUE DÉMOCRATIQUE DU CONGO'
    c.font      = Font(bold=True, size=11, color='B71C1C')
    c.alignment = center
    c.fill      = PatternFill('solid', fgColor='FAFAFA')
    ws.row_dimensions[1].height = 20

    # ══════════════════════════════════════════════
    # LIGNE 2 : Ministère + nom école
    # ══════════════════════════════════════════════
    school_nom = school.nom if school else 'ÉCOLE'
    ws.merge_cells(f'A2:{last_col_letter}2')
    c = ws['A2']
    c.value     = f"Ministère de l'EPST  ·  {school_nom}"
    c.font      = Font(bold=True, size=12, color='1565C0')
    c.alignment = center
    c.fill      = PatternFill('solid', fgColor='E3F2FD')
    ws.row_dimensions[2].height = 22

    # ══════════════════════════════════════════════
    # LIGNE 3 : Titre du document
    # ══════════════════════════════════════════════
    ws.merge_cells(f'A3:{last_col_letter}3')
    c = ws['A3']
    c.value     = 'FEUILLE DE NOTES'
    c.font      = Font(bold=True, size=14, color='FFFFFF')
    c.alignment = center
    c.fill      = fill_blue_dark
    ws.row_dimensions[3].height = 26

    # ══════════════════════════════════════════════
    # LIGNE 4 : Infos matière / classe / enseignant
    # ══════════════════════════════════════════════
    try:
        ens_nom = mc.enseignant.nom_complet
    except Exception:
        ens_nom = '—'
    school_ann = str(mc.classe.annee_scolaire) if hasattr(mc.classe, 'annee_scolaire') and mc.classe.annee_scolaire else ''

    mid = nb_cols // 2
    # Bloc gauche : matière & classe
    ws.merge_cells(f'A4:{get_column_letter(mid)}4')
    c = ws['A4']
    c.value     = f"Matière : {mc.matiere.nom}   |   Classe : {mc.classe}"
    c.font      = Font(bold=True, size=10, color='1565C0')
    c.alignment = left
    c.fill      = PatternFill('solid', fgColor='E8F5E9')
    ws.row_dimensions[4].height = 20

    # Bloc droit : enseignant & année
    ws.merge_cells(f'{get_column_letter(mid+1)}4:{last_col_letter}4')
    c = ws[f'{get_column_letter(mid+1)}4']
    c.value     = f"Enseignant : {ens_nom}   |   Année : {school_ann}"
    c.font      = Font(bold=True, size=10, color='1565C0')
    c.alignment = right_al
    c.fill      = PatternFill('solid', fgColor='E8F5E9')

    # ══════════════════════════════════════════════
    # LIGNE 5 : Métadonnées cachées (fond blanc pur, police blanche = invisible)
    # ══════════════════════════════════════════════
    ws['A5'] = 'META_MC_ID'
    ws['B5'] = mc.pk
    ws['A5'].font = Font(color='FFFFFF', size=1)
    ws['B5'].font = Font(color='FFFFFF', size=1)
    ws.row_dimensions[5].height = 4

    # ══════════════════════════════════════════════
    # LIGNE 6 : En-têtes colonnes
    # ══════════════════════════════════════════════
    headers = ['N°', 'Matricule', 'Nom', 'Postnom'] + [label for _, label in periodes]
    for col_i, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=col_i, value=h)
        c.font      = white_bold
        c.fill      = fill_blue_dark
        c.alignment = center
        c.border    = thin_border('90A4AE')
    ws.row_dimensions[6].height = 32

    # ══════════════════════════════════════════════
    # LIGNE 7 : MAXIMA
    # ══════════════════════════════════════════════
    mx_vals = ['', '', 'MAXIMA /', ''] + [
        max_normal * 2 if code in ('EXAM1', 'EXAM2') else max_normal
        for code, _ in periodes
    ]
    for col_i, val in enumerate(mx_vals, 1):
        c = ws.cell(row=7, column=col_i, value=val)
        c.font      = Font(bold=True, size=9, color='1565C0')
        c.fill      = fill_blue_light
        c.alignment = center
        c.border    = thin_border('90A4AE')
    ws.row_dimensions[7].height = 18

    # ── Auto-filter sur ligne 6 ──
    ws.auto_filter.ref = f'A6:{last_col_letter}6'

    # ══════════════════════════════════════════════
    # LIGNES 8+ : Données élèves
    # ══════════════════════════════════════════════
    first_data_row = 8
    for i, eleve in enumerate(eleves):
        row = first_data_row + i
        is_even = (i % 2 == 0)
        row_fill = fill_stripe if is_even else PatternFill('solid', fgColor='FFFFFF')

        # Colonnes fixes : N°, Matricule, Nom, Postnom
        row_data = [i + 1, eleve.matricule, eleve.nom, getattr(eleve, 'postnom', '')]
        for col_i, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col_i, value=val)
            c.fill      = fill_gray_light if is_even else PatternFill('solid', fgColor='FAFAFA')
            c.font      = Font(bold=(col_i == 3), size=9, color='212121')
            c.alignment = center if col_i == 1 else left
            c.border    = thin_border()

        # Colonnes notes
        for col_i, (code, _) in enumerate(periodes):
            col = 5 + col_i
            max_v = max_normal * 2 if code in ('EXAM1', 'EXAM2') else max_normal
            try:
                note = Note.objects.get(eleve=eleve, matiere_classe=mc, periode=code)
                val  = float(note.valeur)
            except Note.DoesNotExist:
                val = None

            c = ws.cell(row=row, column=col, value=val)
            c.alignment = center
            c.border    = thin_border()
            c.font      = Font(bold=bool(val is not None), size=10,
                               color='1565C0' if val is not None else '9E9E9E')
            c.fill = fill_note if val is not None else (fill_stripe if is_even else PatternFill('solid', fgColor='FFFFFF'))

            # Validation données Excel
            dv = DataValidation(
                type='decimal', operator='between',
                formula1='0', formula2=str(max_v),
                showErrorMessage=True,
                errorTitle='Note invalide',
                error=f'La note doit être comprise entre 0 et {max_v}',
                showInputMessage=True,
                promptTitle='Note',
                prompt=f'Saisir une note entre 0 et {max_v}',
            )
            ws.add_data_validation(dv)
            dv.add(c)

        ws.row_dimensions[row].height = 20

    # ══════════════════════════════════════════════
    # LIGNE SIGNATURE (après les élèves + 2 lignes vides)
    # ══════════════════════════════════════════════
    last_eleve_row = first_data_row + len(eleves) - 1
    sig_row = last_eleve_row + 3

    # Séparateur
    ws.merge_cells(f'A{sig_row}:{last_col_letter}{sig_row}')
    ws[f'A{sig_row}'].fill = PatternFill('solid', fgColor='E3F2FD')
    ws.row_dimensions[sig_row].height = 6

    sig_row += 1
    # Signature enseignant (gauche) + visa préfet (droite)
    ws.merge_cells(f'A{sig_row}:{get_column_letter(mid)}{sig_row}')
    c = ws[f'A{sig_row}']
    c.value     = "Signature de l'enseignant :"
    c.font      = Font(bold=True, size=9, color='424242')
    c.alignment = left

    ws.merge_cells(f'{get_column_letter(mid+1)}{sig_row}:{last_col_letter}{sig_row}')
    c = ws[f'{get_column_letter(mid+1)}{sig_row}']
    c.value     = "Visa du Préfet des études :"
    c.font      = Font(bold=True, size=9, color='424242')
    c.alignment = right_al
    ws.row_dimensions[sig_row].height = 18

    sig_row += 3
    ws.merge_cells(f'A{sig_row}:{get_column_letter(mid)}{sig_row}')
    ws[f'A{sig_row}'].border = Border(bottom=Side(style='medium', color='424242'))

    ws.merge_cells(f'{get_column_letter(mid+1)}{sig_row}:{last_col_letter}{sig_row}')
    ws[f'{get_column_letter(mid+1)}{sig_row}'].border = Border(bottom=Side(style='medium', color='424242'))
    ws.row_dimensions[sig_row].height = 22

    sig_row += 1
    ws.merge_cells(f'A{sig_row}:{get_column_letter(mid)}{sig_row}')
    c = ws[f'A{sig_row}']
    c.value     = "À _________________, le _________________ 20___"
    c.font      = muted_font
    c.alignment = left

    # ── Largeurs colonnes ──
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    for i in range(len(periodes)):
        ltr = get_column_letter(5 + i)
        ws.column_dimensions[ltr].width = 14

    # ── Figer les volets (ligne 8, col 5) ──
    ws.freeze_panes = ws.cell(row=first_data_row, column=5)

    # ── Mise en page : A4 paysage, ajuster à 1 page de large ──
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToPage   = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.6,
                                  header=0.3, footer=0.3)
    # Répéter les lignes d'en-tête à chaque page
    ws.print_title_rows = '1:7'
    # Zone d'impression
    ws.print_area = f'A1:{last_col_letter}{sig_row}'

    # ── Zoom à 90% ──
    ws.sheet_view.zoomScale = 90

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = (
        f"notes_{mc.matiere.nom}_{mc.classe}_{'-'.join(c for c, _ in periodes)}.xlsx"
    ).replace(' ', '_').replace('/', '-')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─── Import Excel — Upload ────────────────────────────────────────────────────
@login_required
def import_notes_excel(request):
    """Étape 1 : upload du fichier Excel exporté."""
    user = request.user
    try:
        teacher = user.teacher_profile
        matieres_classes = MatiereClasse.objects.filter(
            enseignant=teacher
        ).select_related('matiere', 'classe') if not user.is_prefet() else MatiereClasse.objects.select_related('matiere', 'classe')
    except Exception:
        matieres_classes = MatiereClasse.objects.none()

    if user.is_prefet():
        matieres_classes = MatiereClasse.objects.select_related('matiere', 'classe', 'classe__section').order_by('classe__nom', 'matiere__nom')

    if request.method == 'POST' and request.FILES.get('fichier_excel'):
        import openpyxl, io
        fichier = request.FILES['fichier_excel']
        try:
            wb = openpyxl.load_workbook(io.BytesIO(fichier.read()), data_only=True)
            ws = wb.active

            # Lire métadonnées
            mc_id = ws['B5'].value
            if not mc_id:
                messages.error(request, "Fichier invalide : métadonnées manquantes. Utilisez uniquement les fichiers exportés depuis ce système.")
                return redirect('import_notes_excel')

            mc = get_object_or_404(MatiereClasse, pk=mc_id)

            # Sécurité
            if not user.is_prefet():
                try:
                    if mc.enseignant != user.teacher_profile:
                        messages.error(request, "Ce fichier ne vous appartient pas.")
                        return redirect('import_notes_excel')
                except Exception:
                    return redirect('import_notes_excel')

            # Lire en-têtes (ligne 6) pour détecter les périodes
            header_row = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
            periode_labels = {
                '1ère Période': '1P', '2ème Période': '2P', 'Examen S1': 'EXAM1',
                '3ème Période': '3P', '4ème Période': '4P', 'Examen S2': 'EXAM2',
                'Repêchage': 'REPECHAGE',
            }
            # Colonnes des périodes (index 4+)
            periode_cols = []
            for col_i, h in enumerate(header_row[4:], 4):
                if h and h in periode_labels:
                    periode_cols.append((col_i, periode_labels[h], h))

            # Lire données (à partir de ligne 8, ligne 7 = MAXIMA)
            all_periodes = [
                ('1P', '1ère P'), ('2P', '2ème P'), ('EXAM1', 'Exam S1'),
                ('3P', '3ème P'), ('4P', '4ème P'), ('EXAM2', 'Exam S2'),
                ('REPECHAGE', 'Repêchage'),
            ]
            eleves = Student.objects.filter(classe=mc.classe).order_by('nom', 'postnom')
            eleve_map = {e.matricule: e for e in eleves}

            preview_data = []
            errors = []

            for row in ws.iter_rows(min_row=8, values_only=True):
                if not row[1]:  # Matricule vide = fin du tableau
                    break
                matricule = str(row[1]).strip()
                eleve = eleve_map.get(matricule)
                if not eleve:
                    errors.append(f"Matricule inconnu : {matricule}")
                    continue

                changes = []
                for col_i, code, label in periode_cols:
                    max_v = mc.matiere.maxima * (2 if code in ('EXAM1', 'EXAM2') else 1)
                    new_val = row[col_i]
                    if new_val is not None:
                        try:
                            new_dec = Decimal(str(new_val)).quantize(Decimal('0.01'))
                        except Exception:
                            errors.append(f"{eleve.nom_complet} / {label} : valeur invalide '{new_val}'")
                            continue
                        if new_dec < 0 or new_dec > max_v:
                            errors.append(f"{eleve.nom_complet} / {label} : {new_dec} hors limites (0-{max_v})")
                            continue
                    else:
                        new_dec = None

                    # Valeur actuelle en base
                    try:
                        existing = Note.objects.get(eleve=eleve, matiere_classe=mc, periode=code)
                        old_dec = existing.valeur
                    except Note.DoesNotExist:
                        old_dec = None

                    changed = (new_dec != old_dec)
                    changes.append({
                        'periode': code, 'label': label,
                        'avant': old_dec, 'apres': new_dec,
                        'changed': changed,
                    })

                preview_data.append({
                    'eleve': eleve,
                    'changes': changes,
                    'has_changes': any(c['changed'] for c in changes),
                })

            if errors:
                for err in errors[:5]:
                    messages.warning(request, err)

            # Stocker en session pour confirmation
            session_data = []
            for row in preview_data:
                session_data.append({
                    'eleve_id': row['eleve'].pk,
                    'eleve_nom': row['eleve'].nom_complet,
                    'changes': [
                        {
                            'periode': c['periode'], 'label': c['label'],
                            'avant': str(c['avant']) if c['avant'] is not None else None,
                            'apres': str(c['apres']) if c['apres'] is not None else None,
                            'changed': c['changed'],
                        }
                        for c in row['changes']
                    ],
                    'has_changes': row['has_changes'],
                })
            request.session['import_preview'] = session_data
            request.session['import_mc_id']   = mc.pk

            return render(request, 'grades/import_preview.html', {
                'mc': mc,
                'preview_data': preview_data,
                'errors': errors,
                'nb_changes': sum(1 for r in preview_data if r['has_changes']),
            })

        except Exception as e:
            messages.error(request, f"Erreur lecture fichier : {e}")
            return redirect('import_notes_excel')

    return render(request, 'grades/import_notes.html', {'matieres_classes': matieres_classes})


# ─── Import Excel — Confirmation ──────────────────────────────────────────────
@login_required
@require_POST
def import_notes_preview(request):
    """Étape 2 : confirmation et enregistrement des notes importées."""
    from django.db import transaction
    session_data = request.session.get('import_preview')
    mc_id        = request.session.get('import_mc_id')

    if not session_data or not mc_id:
        messages.error(request, "Session expirée. Veuillez réimporter le fichier.")
        return redirect('import_notes_excel')

    mc = get_object_or_404(MatiereClasse, pk=mc_id)

    # Sécurité
    user = request.user
    if not user.is_prefet():
        try:
            if mc.enseignant != user.teacher_profile:
                messages.error(request, "Non autorisé.")
                return redirect('saisie_notes')
        except Exception:
            return redirect('saisie_notes')

    # Pré-charger tous les élèves concernés pour éviter les requêtes N+1
    eleve_ids = [row['eleve_id'] for row in session_data if row.get('has_changes')]
    eleves_map = {e.pk: e for e in Student.objects.filter(pk__in=eleve_ids)}

    nb_saved = 0
    try:
        with transaction.atomic():
            for row in session_data:
                if not row['has_changes']:
                    continue
                eleve = eleves_map.get(row['eleve_id'])
                if eleve is None:
                    continue
                for change in row['changes']:
                    if not change['changed']:
                        continue
                    new_val = change['apres']
                    if new_val is None:
                        Note.objects.filter(eleve=eleve, matiere_classe=mc, periode=change['periode']).delete()
                    else:
                        Note.objects.update_or_create(
                            eleve=eleve, matiere_classe=mc, periode=change['periode'],
                            defaults={'valeur': Decimal(new_val)}
                        )
                    nb_saved += 1
    except Exception as e:
        messages.error(request, "Une erreur est survenue lors de l'importation. Aucune note n'a été modifiée.")
        return redirect('import_notes_excel')

    # Nettoyer la session
    del request.session['import_preview']
    del request.session['import_mc_id']

    messages.success(request, f"{nb_saved} note(s) importée(s) avec succès pour {mc.matiere} — {mc.classe}.")
    return redirect(f'/notes/?mc={mc.pk}&periode=1P')
